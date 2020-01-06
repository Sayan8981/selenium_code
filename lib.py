import os
from selenium.webdriver.common.by import By
import openpyxl

class common_lib:

    def browser_open(self,driver,request_url):       
        driver.get(request_url)
        driver.set_window_size(1800,900)
        return driver

    def find_element_by_name(self,browser,name):
        search_query=browser.find_element_by_name(name)
        print("name tag selected", search_query.is_displayed()) 
        return search_query   
    
    def find_element_by_id(self,browser,id):
        #import pdb;pdb.set_trace()
        search_query=browser.find_element_by_id(id)
        print("id tag selected", search_query.is_displayed()) 
        return search_query

    def find_element_by_link_text(self,browser,link):
        #import pdb;pdb.set_trace()
        search_query=browser.find_element_by_link_text(link)
        print("link selected", search_query.is_displayed()) 
        return search_query  

    def find_element_by_xpath(self,browser,xpath):
        #import pdb;pdb.set_trace()
        search_query=browser.find_element_by_xpath(xpath)
        print("xpath element selected", search_query.is_displayed()) 
        return search_query

    def find_element_by_class(self,browser,class_name):
        search_query=browser.find_elements(By.CLASS_NAME,class_name)
        return search_query 

    def find_elements_by_tag_name(self,browser,tag_name):
        search_query=browser.find_elements_by_tag_name(tag_name)                
        return search_query

    def scroll_down_page_till_end(self,browser):
        return browser.execute_script("window.scrollBy(0,document.body.scrollHeight)")

    def scroll_down_by_element(self,browser,element_to_scroll):
        return browser.execute_script("arguments[0].scrollIntoView();",element_to_scroll)        

    def browser_close(self,browser):
        #time.sleep(250)
        browser.close()  

    def excel_file_action(self,file):
        #import pdb;pdb.set_trace()
        file_sheet = openpyxl.load_workbook(os.getcwd()+file,data_only=True)
        return file_sheet  #file.get_sheet_by_name('sheet1') 

    def create_excel_file(self,filepath):
        wb = openpyxl.Workbook()
        wb.save(filename = filepath)
        return wb  

    def get_rowcount_from_sheet(self,file,sheetName):
        workbook=openpyxl.load_workbook(file)
        sheet=workbook.get_sheet_by_name(sheetName)
        return sheet.max_row

    def get_columncount_from_sheet(self,file,sheetName):
        workbook=openpyxl.load_workbook(file)
        sheet=workbook.get_sheet_by_name(sheetName)
        return sheet.max_column

    def read_sheet_data(self,file,sheetName,rownum,columnno):
        workbook=openpyxl.load_workbook(file)
        sheet=workbook.get_sheet_by_name(sheetName)
        return sheet.cell(row=rownum,column=columnno).value 

    def write_data(self,file,sheetName,rownum,columnno,data):
        workbook=openpyxl.load_workbook(file)
        sheet=workbook.get_sheet_by_name(sheetName)
        sheet.cell(row=rownum,column=columnno).value=data 
        workbook.save(file)              
         

